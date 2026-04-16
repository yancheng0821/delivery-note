"""Tests for generator module."""

import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from io import BytesIO
from generator import generate_delivery_note, split_amount_digits

TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "..", "templates", "送货单(三联针式打印)1.xlsx"
)

CUSTOMER = {
    "name": "测试客户公司",
    "address": "武汉市测试路1号",
    "contact": "张三",
    "phone": "13800138000",
}


def test_split_amount_digits_basic():
    # 360.00 -> [None, None, None, 3, 6, 0, 0, 0]
    assert split_amount_digits(360.00) == [None, None, None, 3, 6, 0, 0, 0]


def test_split_amount_digits_with_decimals():
    # 12345.67 -> [None, 1, 2, 3, 4, 5, 6, 7]
    assert split_amount_digits(12345.67) == [None, 1, 2, 3, 4, 5, 6, 7]


def test_split_amount_digits_zero():
    assert split_amount_digits(0) == [None, None, None, None, None, 0, 0, 0]


def test_split_amount_digits_large():
    # 99999.99 -> [None, 9, 9, 9, 9, 9, 9, 9]
    assert split_amount_digits(99999.99) == [None, 9, 9, 9, 9, 9, 9, 9]


def test_generate_returns_bytes():
    materials = [
        {
            "material_name": "磁力启动器",
            "spec": "【QCX5-22】5.5-7.5KW_380V空压机",
            "quantity": 3,
            "unit": "套",
            "unit_price": 120,
        }
    ]
    result = generate_delivery_note(CUSTOMER, materials, TEMPLATE_PATH)
    assert isinstance(result, BytesIO)
    wb = openpyxl.load_workbook(result)
    ws = wb.active
    assert ws is not None
    wb.close()


def test_generate_fills_customer_info():
    materials = [
        {
            "material_name": "磁力启动器",
            "spec": "型号A",
            "quantity": 1,
            "unit": "套",
            "unit_price": 100,
        }
    ]
    result = generate_delivery_note(CUSTOMER, materials, TEMPLATE_PATH)
    wb = openpyxl.load_workbook(result)
    ws = wb.active
    assert "测试客户公司" in str(ws["B5"].value)
    assert "编号" in str(ws["E5"].value)
    b6 = str(ws["B6"].value)
    assert "武汉市测试路1号" in b6
    assert "张三" in b6
    assert "13800138000" in b6
    wb.close()


def test_generate_fills_item_row():
    materials = [
        {
            "material_name": "磁力启动器",
            "spec": "型号A",
            "quantity": 3,
            "unit": "套",
            "unit_price": 120,
        }
    ]
    result = generate_delivery_note(CUSTOMER, materials, TEMPLATE_PATH)
    wb = openpyxl.load_workbook(result)
    ws = wb.active
    assert ws["B9"].value == 1
    assert "磁力启动器" in str(ws["C9"].value)
    assert ws["D9"].value == 3       # 数量
    assert ws["E9"].value == "套"    # 单位
    assert ws["F9"].value == 120     # 单价
    assert ws["G9"].value == 360     # 总价
    # Amount 360.00 -> digits [None,None,None,3,6,0,0,0] starting col H
    assert ws["K9"].value == 3   # 佰
    assert ws["L9"].value == 6   # 拾
    assert ws["M9"].value == 0   # 元
    wb.close()


def test_generate_all_items_in_one_note():
    """All items go into a single delivery note without pagination."""
    materials = [
        {
            "material_name": f"物料{i}",
            "spec": f"规格{i}",
            "quantity": 1,
            "unit": "个",
            "unit_price": 100,
        }
        for i in range(1, 9)
    ]
    result = generate_delivery_note(CUSTOMER, materials, TEMPLATE_PATH)
    wb = openpyxl.load_workbook(result)
    ws = wb.active
    # All 8 items appear sequentially starting at row 9
    assert ws["B9"].value == 1
    assert "物料1" in str(ws["C9"].value)
    assert ws["B14"].value == 6
    assert "物料6" in str(ws["C14"].value)
    assert ws["B15"].value == 7
    assert "物料7" in str(ws["C15"].value)
    assert ws["B16"].value == 8
    assert "物料8" in str(ws["C16"].value)
    wb.close()
