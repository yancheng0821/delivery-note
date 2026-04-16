"""Generate delivery note Excel files from template."""

from copy import copy
from datetime import date
from io import BytesIO

import openpyxl


def split_amount_digits(amount: float) -> list:
    """Split an amount into 8 digit positions: 拾万,万,仟,佰,拾,元,角,分.

    Returns a list of 8 elements. Leading zeros are None, trailing zeros are 0.
    """
    amount = round(amount, 2)

    int_part = int(amount)
    dec_part = round((amount - int_part) * 100)

    jiao = dec_part // 10
    fen = dec_part % 10

    digits = []
    for divisor in [100000, 10000, 1000, 100, 10, 1]:
        digits.append(int_part // divisor)
        int_part = int_part % divisor

    digits.append(jiao)
    digits.append(fen)

    # Replace leading zeros with None (keep 元 position at index 5)
    result = []
    leading = True
    for i, d in enumerate(digits):
        if leading and d == 0 and i < 5:
            result.append(None)
        else:
            leading = False
            result.append(d)

    return result


# Template row offsets (relative to copy_start_row)
COPY_START_ROW = 2
DATE_ROW_OFFSET = 2       # row 4
CUSTOMER_ROW_OFFSET = 3   # row 5
ADDRESS_ROW_OFFSET = 4    # row 6
ITEM_START_OFFSET = 7     # row 9
ITEM_COUNT = 6            # item rows in template per copy


def _copy_row_style(ws, src_row: int, dst_row: int):
    """Copy style from src_row to dst_row."""
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)


def _insert_extra_item_rows(ws, copy_start_row: int, extra_count: int):
    """Insert extra item rows after the template's 6 rows, copying style."""
    if extra_count <= 0:
        return
    # Last template item row is the style source
    src_row = copy_start_row + ITEM_START_OFFSET + ITEM_COUNT - 1
    insert_at = copy_start_row + ITEM_START_OFFSET + ITEM_COUNT
    for i in range(extra_count):
        ws.insert_rows(insert_at + i)
        _copy_row_style(ws, src_row, insert_at + i)


def _fill_copy(ws, copy_start_row: int, customer: dict, materials: list[dict],
               order_number: str, today: date):
    """Fill one copy of the delivery note with all materials."""
    date_str = f"                                 {today.year}      {today.month:02d}      {today.day:02d}日"
    ws.cell(row=copy_start_row + DATE_ROW_OFFSET, column=2).value = date_str

    ws.cell(row=copy_start_row + CUSTOMER_ROW_OFFSET, column=2).value = f"客户名称：{customer['name']}"
    ws.cell(row=copy_start_row + CUSTOMER_ROW_OFFSET, column=5).value = f"                    编号：{order_number}"

    addr = customer.get("address", "")
    contact = customer.get("contact", "")
    phone = customer.get("phone", "")
    ws.cell(row=copy_start_row + ADDRESS_ROW_OFFSET, column=2).value = (
        f"发往地址：{addr}                                             联系人：{contact}      电话：{phone}"
    )

    n = len(materials)
    # Always iterate at least ITEM_COUNT rows to clear unused template rows
    rows_to_iterate = max(n, ITEM_COUNT)

    page_total = 0.0
    for i in range(rows_to_iterate):
        row = copy_start_row + ITEM_START_OFFSET + i
        if i < n:
            mat = materials[i]
            amount = round(mat["unit_price"] * mat["quantity"], 2)
            page_total += amount

            ws.cell(row=row, column=2).value = i + 1
            name_spec = mat["material_name"]
            if mat.get("spec"):
                name_spec += " " + mat["spec"]
            ws.cell(row=row, column=3).value = name_spec
            ws.cell(row=row, column=4).value = int(mat["quantity"])  # 数量
            ws.cell(row=row, column=5).value = mat["unit"]           # 单位
            ws.cell(row=row, column=6).value = mat["unit_price"]     # 单价
            ws.cell(row=row, column=7).value = amount                # 总价

            digits = split_amount_digits(amount)
            for col_idx, digit in enumerate(digits):
                ws.cell(row=row, column=8 + col_idx).value = digit   # 金额拆位 H-O

            ws.cell(row=row, column=16).value = mat.get("remark", "") or ""  # 备注
        else:
            for col in [2, 3, 4, 5, 6, 7]:
                ws.cell(row=row, column=col).value = None
            for col_idx in range(8):
                ws.cell(row=row, column=8 + col_idx).value = None
            ws.cell(row=row, column=16).value = None

    # Total row shifts down by any extra rows inserted
    total_row = copy_start_row + ITEM_START_OFFSET + rows_to_iterate
    total_digits = split_amount_digits(page_total)
    for col_idx, digit in enumerate(total_digits):
        ws.cell(row=total_row, column=8 + col_idx).value = digit


def generate_delivery_note(customer: dict, materials: list[dict],
                           template_path: str) -> BytesIO:
    """Generate a delivery note Excel from template.

    All materials are placed in a single copy (no pagination).
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    today = date.today()
    order_number = today.strftime('%Y%m%d') + "-001"
    extra = max(0, len(materials) - ITEM_COUNT)

    _insert_extra_item_rows(ws, COPY_START_ROW, extra)
    _fill_copy(ws, COPY_START_ROW, customer, materials, order_number, today)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    return output
