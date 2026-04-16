"""Parse material purchase Excel files."""

import openpyxl


def parse_material_excel(file_path: str) -> list[dict]:
    """Parse a material purchase Excel file and return a list of material dicts.

    Expected format (row 1 = header, data from row 2):
      A: 物料名称, B: 规格/型号, C: 数量, D: 单位, E: 单价, F: 金额, G: 备注
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    materials = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Column A = material_name (index 0)
        material_name = row[0].value
        if not material_name or str(material_name).strip() == "":
            continue

        material_name = str(material_name).strip()

        spec = row[1].value
        spec = str(spec).strip() if spec else ""

        raw_qty = row[2].value   # Column C
        raw_price = row[4].value  # Column E
        unit = row[3].value       # Column D

        try:
            quantity = float(raw_qty) if raw_qty is not None else 0.0
        except (ValueError, TypeError):
            quantity = 0.0

        try:
            unit_price = float(raw_price) if raw_price is not None else 0.0
        except (ValueError, TypeError):
            unit_price = 0.0

        amount = round(unit_price * quantity, 2)

        remark = row[6].value if len(row) > 6 else None  # Column G
        remark = str(remark).strip() if remark else ""

        materials.append({
            "index": len(materials) + 1,
            "material_name": material_name,
            "spec": spec,
            "quantity": quantity,
            "unit": str(unit).strip() if unit else "",
            "unit_price": unit_price,
            "amount": amount,
            "remark": remark,
        })

    wb.close()
    return materials
